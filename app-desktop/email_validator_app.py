# -*- coding: utf-8 -*-
"""
Validador de E-mails - Aplicativo Desktop
==========================================
Valida e-mails de planilhas Excel (.xlsx/.xls) ou CSV em dois modos:

  - Rapido   : valida sintaxe (RFC) + registro MX do dominio (via dnspython)
  - Completo : tudo do modo Rapido + verificacao SMTP (handshake com o servidor)

Resultado exportavel para Excel ou CSV com status e detalhes de cada e-mail.

Autor: gerado com Claude Code
"""

import os
import re
import csv
import sys
import time
import socket
import random
import string
import smtplib
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed

# A interface usa tkinter, mas o nucleo de validacao (regex, MX, SMTP, leitura
# e escrita de planilhas) NAO depende dele. Tornamos o import opcional para que
# o modulo possa ser importado como biblioteca num servidor headless (Linux/
# container) sem tkinter instalado. A GUI so e construida em main() (__main__).
try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except ImportError:
    tk = None
    ttk = filedialog = messagebox = None

# --- Dependencias de terceiros (verificadas no arranque) -------------------
try:
    import dns.resolver
except ImportError:
    dns = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ===========================================================================
#  NUCLEO DE VALIDACAO
# ===========================================================================

# Regex de sintaxe pragmatica (cobre a esmagadora maioria dos casos reais).
EMAIL_REGEX = re.compile(
    r"^(?=.{1,254}$)"                       # comprimento total
    r"(?=.{1,64}@)"                         # parte local <= 64
    r"[A-Za-z0-9!#$%&'*+/=?^_`{|}~-]+"
    r"(?:\.[A-Za-z0-9!#$%&'*+/=?^_`{|}~-]+)*"
    r"@"
    r"(?:[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?\.)+"
    r"[A-Za-z]{2,63}$"
)

# Cache de MX por dominio para nao repetir consultas DNS.
_mx_cache = {}
_mx_cache_lock = threading.Lock()

# Remetente e nome de apresentacao (HELO/EHLO) usados no envelope SMTP. NAO
# envia e-mail de verdade (encerra antes do DATA). Em producao (VPS) estes
# devem refletir um dominio REAL com rDNS/SPF validos, senao muitos servidores
# rejeitam ou mentem na resposta. Configuraveis em runtime (o app web seta a
# partir de variaveis de ambiente).
SMTP_FROM = "verify@example.com"
SMTP_HELO_HOST = "localhost"          # FQDN de apresentacao (EHLO)
SMTP_USE_STARTTLS = True              # tenta TLS quando o servidor oferece
SMTP_TIMEOUT = 10   # segundos
DNS_TIMEOUT = 5     # segundos

# Quando o SMTP NAO e conclusivo (timeout / servidor nao responde / sem rDNS),
# o veredito vem como "desconhecido". Com esta flag LIGADA, um e-mail
# corporativo valido por MX nesse caso e MANTIDO (mesmo criterio do gmail, que
# tambem so passa por MX) em vez de descartado. Evita perder lead bom so porque
# o IP da VPS nao consegue conversar com o servidor de destino. So um veredito
# SMTP DEFINITIVO ("invalido", ex.: 550) descarta. O controle de bounce, nesse
# cenario, fica com a lista de supressao. Desligada por padrao (comportamento
# antigo do app desktop); o app web liga via variavel de ambiente.
SMTP_KEEP_INCONCLUSIVE = False

# ---------------------------------------------------------------------------
#  SEGURANCAS DO MODO SMTP (evitar bloqueio / blacklist do IP)
# ---------------------------------------------------------------------------
# Teto de threads no modo completo: SMTP pede ritmo baixo, nao paralelismo alto.
SMTP_MAX_WORKERS = 5
# Intervalo minimo entre duas sondas AO MESMO dominio (segundos).
SMTP_PER_DOMAIN_INTERVAL = 4.0
# Ritmo global padrao: maximo de sondas SMTP por minuto (ajustavel na GUI).
SMTP_DEFAULT_RATE_PER_MIN = 20

# --- VERIFICACAO PROFUNDA (mais lenta, mais conclusiva) --------------------
# Reenvio em greylisting: quando o servidor responde 4xx ("tente mais tarde"),
# espera e tenta de novo, em vez de desistir com "desconhecido".
SMTP_GREYLIST_RETRIES = 2          # nº de reenvios apos um 4xx
SMTP_GREYLIST_DELAY = 60           # segundos de espera entre tentativas

# Provedores de e-mail de consumo que usam "accept-all" (respondem 250 para
# qualquer endereco). Sondar via SMTP nao da resposta confiavel e ainda
# arrisca o seu IP -> nesses dominios pulamos o SMTP e aceitamos pelo MX.
ACCEPT_ALL_DOMAINS = {
    "gmail.com", "googlemail.com",
    "hotmail.com", "hotmail.com.br", "outlook.com", "outlook.com.br",
    "live.com", "live.com.br", "msn.com",
    "yahoo.com", "yahoo.com.br", "ymail.com", "rocketmail.com",
    "icloud.com", "me.com", "mac.com",
    "aol.com", "gmx.com", "gmx.net", "zoho.com", "mail.com",
    "proton.me", "protonmail.com",
    "uol.com.br", "bol.com.br", "terra.com.br", "ig.com.br",
    "globo.com", "globomail.com", "r7.com",
}

# ---------------------------------------------------------------------------
#  REDUCAO DE BOUNCE (offline, gratis) - typos, descartaveis, e-mails de funcao
# ---------------------------------------------------------------------------

# Dominios descartaveis / temporarios: bouncam ou sao lixo -> tratados como
# invalidos para nao sujar a campanha nem a reputacao.
DISPOSABLE_DOMAINS = {
    "mailinator.com", "tempmail.com", "temp-mail.org", "10minutemail.com",
    "guerrillamail.com", "guerrillamail.info", "sharklasers.com",
    "yopmail.com", "trashmail.com", "throwawaymail.com", "getnada.com",
    "maildrop.cc", "mailnesia.com", "dispostable.com", "fakeinbox.com",
    "fakemail.net", "tempinbox.com", "spam4.me", "mintemail.com",
    "mohmal.com", "emailondeck.com", "mailcatch.com", "tempr.email",
    "discard.email", "33mail.com", "anonbox.net", "mvrht.com",
    "spamgourmet.com", "mytemp.email", "burnermail.io", "gmial.com",
}

# E-mails "de funcao" (papel) em 3 niveis, pensado para PROSPECCAO B2B:
#
#   ROLE_TRAP  -> caixas de sistema / RFC. Sao classicas armadilhas de spam;
#                 enviar para elas DESTROI reputacao. Tratadas como INVALIDAS.
#   ROLE_SKIP  -> automaticas / nao monitoradas (noreply, newsletter). Entregam,
#                 mas nao adianta prospectar. Validas, porem ARRISCADAS (removiveis).
#   ROLE_SEND  -> caixas comerciais legitimas (contato@, comercial@, vendas@...).
#                 Em dados de CNPJ costumam ser o UNICO contato da empresa e sao
#                 monitoradas -> mantidas como BOAS (seguro) para B2B.
ROLE_TRAP = {
    "postmaster", "abuse", "hostmaster", "webmaster", "root",
    "mailer-daemon", "mailerdaemon", "noc", "nobody", "spam",
    "devnull", "security",
}
ROLE_SKIP = {
    "noreply", "no-reply", "naoresponda", "nao-responda", "donotreply",
    "donotrespond", "newsletter", "news", "notifications", "notification",
    "bounce", "bounces", "automated", "mailer",
}
ROLE_SEND = {
    "info", "contato", "contact", "vendas", "sales", "sac", "suporte",
    "support", "atendimento", "admin", "administrador", "administrator",
    "financeiro", "comercial", "marketing", "rh", "cobranca", "faturamento",
    "mail", "office", "hello", "ola", "geral", "contabilidade", "juridico",
    "compras", "fornecedores", "ouvidoria", "imprensa",
}


def _norm_set(s):
    """Normaliza prefixos para comparacao (remove - e _)."""
    return {p.replace("-", "").replace("_", "") for p in s}


_ROLE_TRAP_N = _norm_set(ROLE_TRAP)
_ROLE_SKIP_N = _norm_set(ROLE_SKIP)
_ROLE_SEND_N = _norm_set(ROLE_SEND)

# Correcao de erros de digitacao comuns no DOMINIO (typo -> correto).
TYPO_DOMAINS = {
    # gmail
    "gmail.con": "gmail.com", "gmail.co": "gmail.com", "gmail.cm": "gmail.com",
    "gmail.comm": "gmail.com", "gmial.com": "gmail.com", "gmai.com": "gmail.com",
    "gmal.com": "gmail.com", "gnail.com": "gmail.com", "gmail.om": "gmail.com",
    "gmaill.com": "gmail.com", "gmail.cob": "gmail.com", " gmail.com": "gmail.com",
    "g-mail.com": "gmail.com", "googlemail.con": "googlemail.com",
    # hotmail
    "hotmail.con": "hotmail.com", "hotmial.com": "hotmail.com",
    "hotmai.com": "hotmail.com", "hotmal.com": "hotmail.com",
    "hotmail.co": "hotmail.com", "hotmail.cm": "hotmail.com",
    "hotmail.comm": "hotmail.com", "hotmaill.com": "hotmail.com",
    "homail.com": "hotmail.com", "hotmailcom": "hotmail.com",
    "hotmail.con.br": "hotmail.com.br", "hotmail.co.br": "hotmail.com.br",
    # outlook
    "outlook.con": "outlook.com", "outlok.com": "outlook.com",
    "outloo.com": "outlook.com", "outook.com": "outlook.com",
    "outlook.co": "outlook.com", "outlook.cm": "outlook.com",
    "outlook.comm": "outlook.com", "outlookk.com": "outlook.com",
    # yahoo
    "yahoo.con": "yahoo.com", "yaho.com": "yahoo.com", "yahooo.com": "yahoo.com",
    "yhoo.com": "yahoo.com", "yahoo.co": "yahoo.com", "yahoo.cm": "yahoo.com",
    "yahoo.con.br": "yahoo.com.br", "yahoo.co.br": "yahoo.com.br",
    # provedores BR
    "uol.con.br": "uol.com.br", "uol.com": "uol.com.br",
    "bol.con.br": "bol.com.br", "terra.con.br": "terra.com.br",
    "ig.con.br": "ig.com.br",
}

# Correcoes genericas de final de dominio (sufixo errado -> certo).
TYPO_TLD = [
    (".con", ".com"), (".cmo", ".com"), (".vom", ".com"), (".xom", ".com"),
    (".comm", ".com"), (".co.br.br", ".com.br"), (".con.br", ".com.br"),
    (".com.br.br", ".com.br"),
]


def fix_common_typos(email):
    """
    Corrige erros de digitacao comuns no dominio.
    Retorna (email_corrigido, foi_corrigido).
    """
    if not email or "@" not in email:
        return email, False
    local, _, domain = email.partition("@")
    domain = domain.strip().strip(".").lower()
    original_domain = domain

    # 1) dominio inteiro num mapa de typos conhecidos
    if domain in TYPO_DOMAINS:
        domain = TYPO_DOMAINS[domain]
    else:
        # 2) corrige duplos pontos e sufixos errados
        while ".." in domain:
            domain = domain.replace("..", ".")
        for wrong, right in TYPO_TLD:
            if domain.endswith(wrong):
                domain = domain[: -len(wrong)] + right
                break

    novo = f"{local.strip()}@{domain}"
    return novo, (domain != original_domain)


def classify_role(email):
    """
    Classifica a parte local em: 'trap', 'skip', 'send' ou '' (pessoal).
    Pensado para B2B: traps sao removidos, skip sao arriscados, send (caixas
    comerciais) sao mantidos como bons.
    """
    local = email.split("@", 1)[0].lower()
    # normaliza: remove sufixo +tag, pontos, hifens e underscores
    local = local.split("+", 1)[0].replace(".", "").replace("-", "").replace("_", "")
    if local in _ROLE_TRAP_N:
        return "trap"
    if local in _ROLE_SKIP_N:
        return "skip"
    if local in _ROLE_SEND_N:
        return "send"
    return ""


def classify_risk(status, tipo):
    """
    Classifica o risco de envio:
      seguro    -> valido (pessoal OU caixa comercial B2B legitima)
      arriscado -> nao conclusivo (desconhecido) ou automatico (noreply/news)
      invalido  -> nao enviar (inclui traps: postmaster, abuse...)
    """
    if status == "invalido":
        return "invalido"
    if status == "desconhecido":
        return "arriscado"
    if tipo == "skip":
        return "arriscado"
    # tipo 'funcao_b2b' (contato@, comercial@...) e tratado como seguro p/ B2B
    return "seguro"

# --- Estruturas de pacing (limitacao de ritmo), thread-safe ---------------
_smtp_rate_lock = threading.Lock()
_smtp_last_global = [0.0]                 # instante da ultima sonda (qualquer dominio)
_smtp_min_interval = [60.0 / SMTP_DEFAULT_RATE_PER_MIN]   # seg entre sondas (global)

_domain_guard = threading.Lock()
_domain_locks = {}                        # dominio -> Lock (1 sonda por vez por dominio)
_domain_last = {}                         # dominio -> instante da ultima sonda

# Cache de catch-all por dominio: True (aceita-tudo), False (verifica caixas),
# None (inconclusivo). Evita re-sondar o mesmo dominio.
_catchall_cache = {}
_catchall_lock = threading.Lock()


def _random_localpart(n=16):
    """Gera um nome de caixa aleatorio que quase de certeza nao existe."""
    return "".join(random.choice(string.ascii_lowercase + string.digits)
                   for _ in range(n))


# --- Throttle de consultas DNS (grande escala) ------------------------------
# Limita as consultas MX por segundo para nao ser bloqueado pelos resolvers
# publicos (8.8.8.8, 1.1.1.1...) ao processar centenas de milhares de linhas.
# So conta nas consultas REAIS (cache miss); cache hit nao espera.
DNS_DEFAULT_RATE_PER_SEC = 50
_dns_rate_lock = threading.Lock()
_dns_last = [0.0]
_dns_min_interval = [1.0 / DNS_DEFAULT_RATE_PER_SEC]


def set_dns_rate_per_sec(rate):
    """Define o ritmo global de consultas DNS (consultas por segundo)."""
    rate = max(1.0, float(rate))
    _dns_min_interval[0] = 1.0 / rate


def _dns_rate_wait():
    """Bloqueia ate respeitar o ritmo global entre consultas DNS reais."""
    with _dns_rate_lock:
        now = time.monotonic()
        wait = _dns_last[0] + _dns_min_interval[0] - now
        if wait > 0:
            time.sleep(wait)
            now = time.monotonic()
        _dns_last[0] = now


def set_smtp_rate_per_min(rate):
    """Define o ritmo global de sondas SMTP (sondas por minuto)."""
    rate = max(1, int(rate))
    _smtp_min_interval[0] = 60.0 / rate


def _smtp_rate_wait():
    """Bloqueia ate respeitar o ritmo global entre sondas SMTP."""
    with _smtp_rate_lock:
        now = time.monotonic()
        wait = _smtp_last_global[0] + _smtp_min_interval[0] - now
        if wait > 0:
            time.sleep(wait)
            now = time.monotonic()
        _smtp_last_global[0] = now


def _domain_lock(domain):
    with _domain_guard:
        lk = _domain_locks.get(domain)
        if lk is None:
            lk = threading.Lock()
            _domain_locks[domain] = lk
        return lk

# Servidores DNS publicos usados como fallback. Sao essenciais para o .exe:
# quando empacotado pelo PyInstaller no Windows, o dnspython muitas vezes nao
# consegue ler a configuracao de DNS do sistema e fica SEM nameservers, o que
# faria todas as consultas falharem (e todos os e-mails virarem "sem MX").
PUBLIC_DNS = ["8.8.8.8", "1.1.1.1", "8.8.4.4", "1.0.0.1", "9.9.9.9"]


def _make_resolver():
    """Cria um resolver DNS robusto, com nameservers publicos garantidos."""
    try:
        resolver = dns.resolver.Resolver()          # le config do sistema
    except Exception:
        resolver = dns.resolver.Resolver(configure=False)

    ns = list(resolver.nameservers) if resolver.nameservers else []
    # Garante os publicos como fallback (no fim, depois dos do sistema).
    for pub in PUBLIC_DNS:
        if pub not in ns:
            ns.append(pub)
    resolver.nameservers = ns
    resolver.timeout = DNS_TIMEOUT          # por tentativa a cada servidor
    resolver.lifetime = DNS_TIMEOUT * 3     # orcamento total: permite tentar fallbacks
    return resolver


def check_syntax(email):
    """Retorna True se o e-mail respeita a sintaxe esperada."""
    if not email:
        return False
    return bool(EMAIL_REGEX.match(email.strip()))


def dns_self_test():
    """
    Testa se a resolucao DNS esta a funcionar de todo.
    Retorna (ok: bool, detalhe: str).
    """
    if dns is None:
        return False, "Modulo dnspython nao esta instalado."
    try:
        resolver = _make_resolver()
        resolver.resolve("google.com", "MX")
        return True, "DNS operacional."
    except Exception as exc:
        return False, f"{type(exc).__name__}: {str(exc)[:120]}"


def get_mx_records(domain):
    """
    Devolve (hosts, dns_status) para o dominio.
      hosts      : lista de servidores de e-mail (vazia se nao houver).
      dns_status : "ok"        -> resolveu (com ou sem MX, ver hosts)
                   "nxdomain"  -> dominio NAO existe (inválido de verdade)
                   "no_mx"     -> dominio existe mas nao recebe e-mail
                   "dns_error" -> falha de rede/timeout (NAO conclusivo)
                   "no_dns"    -> dnspython indisponivel
    Usa cache por dominio.
    """
    domain = domain.lower().strip()
    with _mx_cache_lock:
        if domain in _mx_cache:
            return _mx_cache[domain]

    if dns is None:
        result = ([], "no_dns")
        with _mx_cache_lock:
            _mx_cache[domain] = result
        return result

    # cache miss -> consulta real: respeita o ritmo global de DNS
    _dns_rate_wait()

    hosts, status = [], "dns_error"
    try:
        resolver = _make_resolver()
        answers = resolver.resolve(domain, "MX")
        records = sorted(answers, key=lambda r: r.preference)
        hosts = [str(r.exchange).rstrip(".") for r in records]
        # "Null MX" (RFC 7505): um unico MX com troca "." (vira "" apos rstrip)
        # significa que o dominio DECLARA que NAO recebe e-mail -> invalido.
        if any(h == "" for h in hosts) and all(h == "" for h in hosts):
            hosts, status = [], "no_mx"
        else:
            hosts = [h for h in hosts if h]   # descarta trocas vazias
            status = "ok" if hosts else "no_mx"
    except dns.resolver.NXDOMAIN:
        hosts, status = [], "nxdomain"          # dominio nao existe
    except dns.resolver.NoAnswer:
        # Dominio existe mas sem registo MX -> tenta registo A (fallback historico).
        try:
            resolver = _make_resolver()
            resolver.resolve(domain, "A")
            hosts, status = [domain], "ok"
        except dns.resolver.NXDOMAIN:
            hosts, status = [], "nxdomain"
        except dns.resolver.NoAnswer:
            hosts, status = [], "no_mx"
        except Exception:
            hosts, status = [], "dns_error"
    except (dns.resolver.NoNameservers, dns.exception.Timeout,
            dns.resolver.LifetimeTimeout):
        hosts, status = [], "dns_error"         # problema de rede, NAO do e-mail
    except Exception:
        hosts, status = [], "dns_error"

    result = (hosts, status)
    # So cacheia resultados DEFINITIVOS. Um "dns_error" e transitorio (timeout,
    # resolver bloqueado momentaneamente) - cachea-lo faria todos os e-mails
    # seguintes do mesmo dominio cairem como "desconhecido" e serem descartados,
    # mesmo apos a rede voltar. Por isso nao guardamos erros: o proximo e-mail
    # do dominio tenta de novo.
    if status != "dns_error":
        with _mx_cache_lock:
            _mx_cache[domain] = result
    return result


def _pace_domain(domain):
    """Respeita o intervalo por dominio + o ritmo global antes de uma sonda."""
    with _domain_guard:
        last = _domain_last.get(domain, 0.0)
    gap = last + SMTP_PER_DOMAIN_INTERVAL - time.monotonic()
    if gap > 0:
        time.sleep(gap)               # espaca sondas ao mesmo dominio
    _smtp_rate_wait()                 # respeita o ritmo global (sondas/min)
    with _domain_guard:
        _domain_last[domain] = time.monotonic()


def _detect_catch_all_locked(domain, mx_hosts):
    """
    Descobre se o dominio e 'catch-all' (aceita qualquer endereco).
    Deve ser chamado JA com o lock do dominio adquirido.

    Sonda um endereco aleatorio (que quase de certeza nao existe):
      - se o servidor ACEITA   -> True  (catch-all; nao da pra confiar no RCPT)
      - se REJEITA             -> False (servidor verifica caixas de verdade)
      - se inconclusivo        -> None
    Resultado fica em cache por dominio.
    """
    with _catchall_lock:
        if domain in _catchall_cache:
            return _catchall_cache[domain]

    probe = f"{_random_localpart()}@{domain}"
    _pace_domain(domain)
    status, _detail, _temp = _smtp_probe(probe, mx_hosts)
    if status == "valido":
        result = True
    elif status == "invalido":
        result = False
    else:
        result = None

    with _catchall_lock:
        _catchall_cache[domain] = result
    return result


def smtp_check(email, mx_hosts, domain=None, deep=False,
               retries=None, retry_delay=None):
    """
    Faz handshake SMTP ate ao comando RCPT TO para descobrir se a caixa existe.
    Nao envia nenhuma mensagem (encerra antes do DATA).

    retries/retry_delay: se None, usam as constantes do modulo NO MOMENTO da
    chamada (assim o app web pode ajusta-las em runtime, ex.: 0 reenvios para
    nao travar o lote esperando greylisting).

    Aplica segurancas anti-bloqueio: 1 sonda por vez por dominio, intervalo
    minimo por dominio e ritmo global limitado.

    deep=True (verificacao profunda, mais lenta):
      - deteta dominios 'catch-all' (aceita-tudo) e marca como nao confiaveis;
      - reenvia em greylisting (resposta 4xx) ate 'retries' vezes.

    Retorna (status, detalhe):
      status em {"valido", "invalido", "desconhecido"}
    """
    if not mx_hosts:
        return "invalido", "Sem servidor MX"

    if retries is None:
        retries = SMTP_GREYLIST_RETRIES
    if retry_delay is None:
        retry_delay = SMTP_GREYLIST_DELAY

    if domain is None:
        domain = email.rsplit("@", 1)[-1].lower()

    # --- Seguranca: serializa por dominio + respeita intervalos ------------
    dlock = _domain_lock(domain)
    with dlock:                       # so 1 sonda a este dominio de cada vez
        # Verificacao profunda: o dominio aceita qualquer endereco? Se sim, o
        # RCPT do endereco real nao prova nada (vai bouncar na campanha).
        if deep:
            ca = _detect_catch_all_locked(domain, mx_hosts)
            if ca is True:
                return ("desconhecido",
                        "Dominio aceita-tudo (catch-all): caixa nao confirmavel")

        attempts = (retries + 1) if deep else 1
        last = ("desconhecido", "Falha SMTP")
        for i in range(attempts):
            _pace_domain(domain)
            status, detalhe, temporario = _smtp_probe(email, mx_hosts)
            last = (status, detalhe)
            # so vale reenviar se foi recusa temporaria (greylisting) no modo deep
            if not (temporario and deep) or i == attempts - 1:
                break
            time.sleep(retry_delay)   # espera o servidor liberar (greylisting)
        return last


def _smtp_probe(email, mx_hosts):
    """
    Faz o handshake SMTP propriamente dito (sem pacing).
    Retorna (status, detalhe, temporario):
      temporario=True indica recusa 4xx (greylisting) -> vale a pena reenviar.
    """
    last_detail = "Falha SMTP"
    last_temp = False
    for host in mx_hosts:
        server = None
        try:
            server = smtplib.SMTP(timeout=SMTP_TIMEOUT)
            server.connect(host, 25)
            # EHLO com um FQDN real (configuravel). Servidores serios rejeitam
            # ou penalizam HELO generico tipo "example.com".
            server.ehlo(SMTP_HELO_HOST)
            # STARTTLS quando o servidor oferece: alguns so respondem direito
            # ao RCPT sob TLS, e melhora a aceitacao da sonda.
            if SMTP_USE_STARTTLS and server.has_extn("starttls"):
                try:
                    server.starttls()
                    server.ehlo(SMTP_HELO_HOST)
                except Exception:
                    pass   # sem TLS, segue em texto puro
            server.mail(SMTP_FROM)
            code, message = server.rcpt(email)
            try:
                server.quit()
            except Exception:
                pass

            if isinstance(message, bytes):
                message = message.decode("utf-8", "ignore")

            if code in (250, 251):
                return "valido", f"SMTP {code}", False
            if code in (550, 551, 553, 554):
                return "invalido", f"SMTP {code}: {message[:80]}", False
            if code in (450, 451, 452, 421):
                # Greylisting / temporario -> nao conclusivo, mas reenviavel
                last_detail = f"SMTP {code} (temporario): {message[:60]}"
                last_temp = True
                continue
            last_detail = f"SMTP {code}: {message[:60]}"
        except (socket.timeout, smtplib.SMTPServerDisconnected):
            last_detail = "Timeout / desconexao SMTP"
            continue
        except (smtplib.SMTPConnectError, ConnectionRefusedError, OSError):
            last_detail = "Conexao recusada (porta 25 bloqueada?)"
            continue
        except Exception as exc:
            last_detail = f"Erro SMTP: {str(exc)[:60]}"
            continue
        finally:
            if server is not None:
                try:
                    server.close()
                except Exception:
                    pass

    return "desconhecido", last_detail, last_temp


def _finish(result, status, motivo, tipo=None):
    """Preenche status/motivo/risco e devolve o result."""
    result["status"] = status
    result["motivo"] = motivo
    if tipo is not None:
        result["tipo"] = tipo
    result["risco"] = classify_risk(status, result.get("tipo", ""))
    return result


def validate_email(email, mode, deep=False):
    """
    Valida um unico e-mail, com camadas de reducao de bounce.
    mode in {"rapido", "completo"}.
    deep=True ativa a verificacao profunda no SMTP (catch-all + greylisting).
    Retorna dict: email (original), email_final (corrigido), corrigido,
                  status, risco, tipo, motivo.
    """
    raw = "" if email is None else str(email).strip()
    result = {
        "email": raw, "email_final": raw, "corrigido": False,
        "status": "invalido", "risco": "invalido", "tipo": "", "motivo": "",
    }

    # 0) Correcao de erros de digitacao no dominio (recupera leads)
    fixed, corrigido = fix_common_typos(raw)
    result["email_final"] = fixed
    result["corrigido"] = corrigido
    nota_typo = f" [corrigido de {raw}]" if corrigido else ""

    # 1) Sintaxe (sobre o e-mail ja corrigido)
    if not check_syntax(fixed):
        return _finish(result, "invalido", "Sintaxe invalida" + nota_typo)

    domain = fixed.rsplit("@", 1)[1].lower()

    # 1b) Dominio descartavel / temporario -> invalido (bounce/lixo)
    if domain in DISPOSABLE_DOMAINS:
        return _finish(result, "invalido", "Dominio descartavel" + nota_typo,
                       tipo="descartavel")

    # 1c) E-mail de funcao em 3 niveis (B2B)
    role = classify_role(fixed)
    if role == "trap":
        return _finish(
            result, "invalido",
            "Caixa de sistema / armadilha de spam (postmaster, abuse...) - "
            "nao enviar" + nota_typo, tipo="trap")
    if role == "skip":
        result["tipo"] = "skip"          # noreply/newsletter: entrega, mas nao prospecta
    elif role == "send":
        result["tipo"] = "funcao_b2b"    # contato@/comercial@: contato B2B legitimo

    # 2) MX
    mx_hosts, dns_status = get_mx_records(domain)

    if dns_status == "dns_error":
        return _finish(result, "desconhecido",
                       "Falha de DNS (sem conexao / timeout)" + nota_typo)
    if dns_status == "no_dns":
        return _finish(result, "desconhecido", "dnspython indisponivel" + nota_typo)
    if dns_status == "nxdomain":
        return _finish(result, "invalido", "Dominio nao existe" + nota_typo)
    if dns_status == "no_mx" or not mx_hosts:
        return _finish(result, "invalido", "Dominio sem registro MX" + nota_typo)

    base_ok = "Sintaxe OK + MX encontrado"
    if result["tipo"] == "funcao_b2b":
        base_ok += " (caixa comercial B2B)"
    elif result["tipo"] == "skip":
        base_ok += " (automatico: noreply/newsletter)"

    if mode == "rapido":
        return _finish(result, "valido", base_ok + nota_typo)

    # 3) SMTP (modo completo)
    # Seguranca: provedores "accept-all" (Gmail, Outlook, Yahoo...) nao dao
    # resposta SMTP confiavel e sondar so arrisca o IP -> aceita pelo MX.
    if domain in ACCEPT_ALL_DOMAINS:
        return _finish(result, "valido",
                       "Provedor accept-all (MX OK; SMTP ignorado)" + nota_typo)

    status, detalhe = smtp_check(fixed, mx_hosts, domain, deep=deep)
    # SMTP nao conclusivo (timeout / servidor mudo / IP sem rDNS): nao da pra
    # provar que a caixa NAO existe. Mantemos como valido por MX (igual ao
    # gmail) em vez de descartar lead bom. So o veredito DEFINITIVO "invalido"
    # (550 etc.) descarta de verdade.
    if status == "desconhecido" and SMTP_KEEP_INCONCLUSIVE:
        return _finish(result, "valido",
                       "MX OK; SMTP nao conclusivo - mantido" + nota_typo)
    return _finish(result, status, detalhe + nota_typo)


# ===========================================================================
#  LEITURA / ESCRITA DE FICHEIROS
# ===========================================================================

def _read_raw_rows(path):
    """Le todas as linhas de um CSV/Excel como matriz de strings."""
    ext = os.path.splitext(path)[1].lower()
    rows = []

    if ext in (".csv", ".txt"):
        # Deteta delimitador
        with open(path, "r", encoding="utf-8-sig", errors="ignore", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ";" if sample.count(";") > sample.count(",") else ","
            reader = csv.reader(f, delimiter=delimiter)
            rows = [r for r in reader]

    elif ext == ".xls":
        raise RuntimeError(
            "Formato .xls (Excel antigo) nao e suportado. Abra no Excel e "
            "salve como .xlsx, depois tente de novo.")
    elif ext in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise RuntimeError("openpyxl nao esta instalado para ler Excel.")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in r])
        wb.close()
    else:
        raise RuntimeError(f"Formato nao suportado: {ext}")

    return rows


def read_table_from_file(path, column=None):
    """
    Le um ficheiro CSV/Excel preservando TODA a tabela original.
    Deteta (ou recebe) a coluna de e-mails.

    Retorna dict:
      {
        "header": list | None,     # cabecalho (None se nao houver)
        "rows":   list[list],      # apenas as linhas de dados (sem cabecalho)
        "col_idx": int,            # indice da coluna de e-mails
        "emails": list[str],       # e-mails extraidos (na ordem das linhas)
      }
    """
    rows = _read_raw_rows(path)
    if not rows:
        return {"header": None, "rows": [], "col_idx": 0, "emails": []}

    first = rows[0]
    # Decide se a primeira linha e cabecalho (nao contem '@').
    header_has_at = any("@" in str(c) for c in first)
    if header_has_at:
        header = None
        data_rows = rows
    else:
        header = first
        data_rows = rows[1:]

    col_idx = None
    if column is not None and column != "":
        # Pode ser indice numerico ou nome de cabecalho
        if isinstance(column, int):
            col_idx = column
        elif str(column).isdigit():
            col_idx = int(column)
        elif header is not None:
            for i, name in enumerate(header):
                if str(name).strip().lower() == str(column).strip().lower():
                    col_idx = i
                    break

    if col_idx is None and header is not None:
        # tenta cabecalho com 'mail'
        for i, name in enumerate(header):
            if "mail" in str(name).lower():
                col_idx = i
                break

    if col_idx is None:
        # deteta pela coluna que mais contem '@'
        best_i, best_count = None, 0
        sample = data_rows[:200] if len(data_rows) > 200 else data_rows
        ncols = max((len(r) for r in sample), default=0)
        for i in range(ncols):
            count = sum(1 for r in sample if i < len(r) and "@" in str(r[i]))
            if count > best_count:
                best_count, best_i = count, i
        col_idx = best_i if best_i is not None else 0

    emails = []
    for r in data_rows:
        if col_idx < len(r):
            val = str(r[col_idx]).strip()
            if val:
                emails.append(val)

    return {"header": header, "rows": data_rows, "col_idx": col_idx,
            "emails": emails}


def read_emails_from_file(path, column=None):
    """Compatibilidade: devolve apenas a lista de e-mails."""
    return read_table_from_file(path, column)["emails"]


# Caracteres que, no inicio de uma celula, fazem o Excel/LibreOffice tratar o
# conteudo como FORMULA ao abrir o arquivo (CSV/Excel injection, CWE-1236).
# Como as planilhas de leads vem de fontes nao confiaveis (scraping/CNPJ) e a
# saida e feita para ser aberta no Excel, neutralizamos prefixando "'".
_CSV_INJECT_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize_cell(value):
    """Neutraliza injecao de formula prefixando aspa simples quando preciso."""
    s = "" if value is None else str(value)
    if s and s[0] in _CSV_INJECT_CHARS:
        return "'" + s
    return s


def _sanitize_row(row):
    return [_sanitize_cell(c) for c in row]


# Teto de linhas de DADOS por arquivo XLSX (o formato suporta 1.048.576 no
# total, incluindo o cabecalho). Acima disto, a saida e dividida em
# _parte2.xlsx, _parte3.xlsx... para nunca gerar um arquivo invalido.
XLSX_MAX_DATA_ROWS = 1_000_000


def write_results(path, results):
    """Grava o relatorio completo em CSV ou Excel conforme a extensao."""
    ext = os.path.splitext(path)[1].lower()
    headers = ["email", "email_corrigido", "corrigido",
               "status", "risco", "tipo", "motivo"]

    def row_of(r):
        return [
            r.get("email", ""),
            r.get("email_final", ""),
            "sim" if r.get("corrigido") else "",
            r.get("status", ""),
            r.get("risco", ""),
            r.get("tipo", ""),
            r.get("motivo", ""),
        ]

    if ext in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise RuntimeError("openpyxl nao esta instalado para gravar Excel.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"
        ws.append(headers)
        for r in results:
            ws.append(_sanitize_row(row_of(r)))
        wb.save(path)
    else:
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(headers)
            for r in results:
                writer.writerow(_sanitize_row(row_of(r)))


def write_clean_spreadsheet(path, table, keep_fn, info_by_email, dedup=False):
    """
    Grava a planilha original 'limpa': mantem todas as colunas e cabecalho,
    conserva apenas as linhas aprovadas por keep_fn e SUBSTITUI o e-mail pela
    versao corrigida (typo).

    table         : dict devolvido por read_table_from_file
    keep_fn       : funcao(result_dict) -> bool (quais linhas manter)
    info_by_email : dict {email_original_normalizado: result_dict}
    dedup         : se True, mantem so a 1a linha de cada e-mail (corrigido)

    Retorna (total_original, total_mantidas).
    """
    header = table.get("header")
    rows = table.get("rows", [])
    col_idx = table.get("col_idx", 0)

    kept = []
    seen = set()
    for r in rows:
        email = ""
        if col_idx < len(r):
            email = str(r[col_idx]).strip()
        info = info_by_email.get(email)
        if info and keep_fn(info):
            final = info.get("email_final", email)
            if dedup:
                key = final.strip().lower()
                if key in seen:
                    continue
                seen.add(key)
            new_r = list(r)
            # usa o e-mail corrigido na saida (pra nao bouncer por typo)
            if col_idx < len(new_r):
                new_r[col_idx] = final
            kept.append(new_r)

    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise RuntimeError("openpyxl nao esta instalado para gravar Excel.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Limpa"
        if header is not None:
            ws.append(_sanitize_row(header))
        for r in kept:
            ws.append(_sanitize_row(r))
        wb.save(path)
    else:
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, delimiter=";")
            if header is not None:
                writer.writerow(_sanitize_row(header))
            for r in kept:
                writer.writerow(_sanitize_row(r))

    return len(rows), len(kept)


# ===========================================================================
#  PROCESSAMENTO EM GRANDE ESCALA (STREAMING) - centenas de milhares de linhas
# ===========================================================================
# Le linha-a-linha (sem segurar tudo na RAM), valida em modo Rapido (MX, com
# cache por dominio) e grava a planilha limpa de forma incremental em CSV.

def open_table_stream(path, column=None, detect_buffer=60):
    """
    Abre um CSV/Excel em modo STREAMING.
    Retorna (header, col_idx, row_iter, approx_total):
      header     : lista | None
      col_idx    : indice da coluna de e-mail
      row_iter   : gerador que devolve cada linha de dados (lista de str)
      approx_total: int | None (estimativa de linhas, se disponivel)
    NAO carrega o ficheiro inteiro em memoria.
    """
    ext = os.path.splitext(path)[1].lower()

    if ext in (".csv", ".txt"):
        # contagem rapida de linhas (varredura de bytes) para estimar o total
        approx_total = 0
        try:
            with open(path, "rb") as _bf:
                for _chunk in iter(lambda: _bf.read(1024 * 1024), b""):
                    approx_total += _chunk.count(b"\n")
            approx_total = max(0, approx_total)  # cabecalho descontado mais abaixo
        except Exception:
            approx_total = None

        f = open(path, "r", encoding="utf-8-sig", errors="ignore", newline="")
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except Exception:
            delimiter = ";" if sample.count(";") > sample.count(",") else ","
        reader = csv.reader(f, delimiter=delimiter)
        raw_iter = reader
        closer = f.close

    elif ext == ".xls":
        raise RuntimeError(
            "Formato .xls (Excel antigo) nao e suportado. Abra no Excel e "
            "salve como .xlsx, depois tente de novo.")
    elif ext in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise RuntimeError("openpyxl nao esta instalado para ler Excel.")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        approx_total = ws.max_row if isinstance(ws.max_row, int) else None
        cell_iter = ws.iter_rows(values_only=True)
        raw_iter = (["" if c is None else str(c) for c in r] for r in cell_iter)
        closer = wb.close
    else:
        raise RuntimeError(f"Formato nao suportado: {ext}")

    # 1) primeira linha -> decide cabecalho
    try:
        first = next(raw_iter)
    except StopIteration:
        closer()
        return None, 0, iter(()), 0
    first = [str(c) for c in first]
    header_has_at = any("@" in c for c in first)
    header = None if header_has_at else first

    # 2) bufferiza algumas linhas de dados para detetar a coluna de e-mail
    buffer = []
    if header_has_at:
        buffer.append(first)
    while len(buffer) < detect_buffer:
        try:
            r = next(raw_iter)
        except StopIteration:
            break
        buffer.append([str(c) for c in r])

    col_idx = _detect_email_column(header, column, buffer)
    if approx_total and header is not None:
        approx_total = max(0, approx_total - 1)

    def gen():
        try:
            for r in buffer:
                yield r
            for r in raw_iter:
                yield [str(c) for c in r]
        finally:
            try:
                closer()
            except Exception:
                pass

    return header, col_idx, gen(), approx_total


def _detect_email_column(header, column, sample_rows):
    """Descobre o indice da coluna de e-mail (por nome, indice ou conteudo)."""
    # 1) coluna informada explicitamente
    if column is not None and column != "":
        if isinstance(column, int):
            return column
        if str(column).isdigit():
            return int(column)
        if header is not None:
            for i, name in enumerate(header):
                if str(name).strip().lower() == str(column).strip().lower():
                    return i
    # 2) cabecalho com 'mail'
    if header is not None:
        for i, name in enumerate(header):
            if "mail" in str(name).lower():
                return i
    # 3) coluna que mais contem '@' na amostra
    best_i, best_count = 0, 0
    ncols = max((len(r) for r in sample_rows), default=0)
    for i in range(ncols):
        count = sum(1 for r in sample_rows if i < len(r) and "@" in str(r[i]))
        if count > best_count:
            best_count, best_i = count, i
    return best_i


def stream_process(in_path, out_path, column, excluir_arriscados,
                   dns_rate_per_sec, progress_cb=None, stop_event=None,
                   every=2000, dedup=False):
    """
    Processa um ficheiro GRANDE em streaming (modo Rapido) e grava a planilha
    limpa (CSV) de forma incremental. Devolve um dict de contadores.

    Mantem todas as colunas originais; remove linhas sem e-mail, invalidas e
    (opcionalmente) as arriscadas; substitui o e-mail pela versao corrigida.
    Se dedup=True, grava so a 1a ocorrencia de cada e-mail corrigido.
    """
    set_dns_rate_per_sec(dns_rate_per_sec)
    header, col_idx, rows, approx_total = open_table_stream(in_path, column)

    c = {"processadas": 0, "com_email": 0, "sem_email": 0,
         "seguro": 0, "arriscado": 0, "invalido": 0,
         "corrigido": 0, "duplicados": 0, "mantidas": 0,
         "total_estimado": approx_total or 0}
    seen = set()

    # Saida em XLSX (modelo oficial, abre certo no Excel) ou CSV, conforme a
    # extensao escolhida. O XLSX usa o modo write_only do openpyxl: grava as
    # linhas em streaming, sem segurar a planilha inteira na RAM.
    ext = os.path.splitext(out_path)[1].lower()
    use_xlsx = ext in (".xlsx", ".xlsm")
    out_files = []
    if use_xlsx:
        if openpyxl is None:
            raise RuntimeError("openpyxl nao esta instalado para gravar Excel.")
        # Grava em streaming e DIVIDE em _parte2.xlsx... ao passar do teto do
        # formato, para nunca gerar um arquivo invalido (>1.048.576 linhas).
        _base, _xext = os.path.splitext(out_path)
        _xs = {"wb": None, "ws": None, "cur": None, "part": 0, "rows": 0}

        def _open_part():
            if _xs["wb"] is not None:
                _xs["wb"].save(_xs["cur"])
            _xs["part"] += 1
            _xs["cur"] = (out_path if _xs["part"] == 1
                          else f"{_base}_parte{_xs['part']}{_xext}")
            _xs["wb"] = openpyxl.Workbook(write_only=True)
            _xs["ws"] = _xs["wb"].create_sheet("Limpa")
            _xs["rows"] = 0
            out_files.append(_xs["cur"])
            if header is not None:                 # cabecalho em cada parte
                _xs["ws"].append(_sanitize_row(header))

        _open_part()

        def _emit(row):
            if _xs["rows"] >= XLSX_MAX_DATA_ROWS:
                _open_part()
            _xs["ws"].append(_sanitize_row(row))
            _xs["rows"] += 1

        def _emit_header():
            pass                # ja escrito dentro de _open_part (em toda parte)

        def _close_out():
            if _xs["wb"] is not None:
                _xs["wb"].save(_xs["cur"])
    else:
        out = open(out_path, "w", encoding="utf-8-sig", newline="")
        writer = csv.writer(out, delimiter=";")
        out_files.append(out_path)

        def _emit(row):
            writer.writerow(_sanitize_row(row))

        def _emit_header():
            if header is not None:
                _emit(header)

        def _close_out():
            out.close()

    try:
        _emit_header()

        for r in rows:
            if stop_event is not None and stop_event.is_set():
                break
            c["processadas"] += 1

            email = ""
            if col_idx < len(r):
                email = str(r[col_idx]).strip()

            if not email:
                c["sem_email"] += 1
            else:
                c["com_email"] += 1
                res = validate_email(email, "rapido")
                risco = res.get("risco", "")
                if risco == "seguro":
                    c["seguro"] += 1
                elif risco == "arriscado":
                    c["arriscado"] += 1
                else:
                    c["invalido"] += 1
                if res.get("corrigido"):
                    c["corrigido"] += 1

                manter = (res.get("status") == "valido")
                if manter and excluir_arriscados and risco == "arriscado":
                    manter = False
                if manter and dedup:
                    key = res.get("email_final", email).strip().lower()
                    if key in seen:
                        manter = False
                        c["duplicados"] += 1
                    else:
                        seen.add(key)
                if manter:
                    new_r = list(r)
                    if col_idx < len(new_r):
                        new_r[col_idx] = res.get("email_final", email)
                    _emit(new_r)
                    c["mantidas"] += 1

            if progress_cb is not None and c["processadas"] % every == 0:
                progress_cb(dict(c))
    finally:
        _close_out()

    c["arquivos"] = list(out_files)
    if progress_cb is not None:
        progress_cb(dict(c))
    return c


# ===========================================================================
#  INTERFACE GRAFICA
# ===========================================================================

# Acima deste nº de linhas, o botao "Validar lista" passa automaticamente para
# o modo streaming (grava CSV limpo direto em disco, sem travar a tela).
BIG_FILE_THRESHOLD = 50000

# Classe-base da janela: tk.Tk quando ha interface; 'object' permite que o
# modulo seja importado headless (servidor) sem tkinter. A GUI so e instanciada
# em main(), entao a base 'object' nunca chega a ser usada nesse cenario.
_GUI_BASE = tk.Tk if tk is not None else object


class EmailValidatorApp(_GUI_BASE):
    def __init__(self):
        super().__init__()
        self.title("Validador de E-mails")
        self.geometry("820x720")
        self.minsize(700, 600)

        self.input_path = tk.StringVar()
        self.column_name = tk.StringVar()
        self.mode = tk.StringVar(value="rapido")
        self.workers = tk.IntVar(value=20)
        self.smtp_rate = tk.IntVar(value=SMTP_DEFAULT_RATE_PER_MIN)
        self.verificacao_profunda = tk.BooleanVar(value=True)
        self.excluir_arriscados = tk.BooleanVar(value=False)
        self.dedup = tk.BooleanVar(value=True)
        self.dns_rate = tk.IntVar(value=DNS_DEFAULT_RATE_PER_SEC)

        self._table = None
        self._big_t0 = 0.0
        self._emails = []
        self._results = []
        self._queue = queue.Queue()
        self._stop_flag = threading.Event()
        self._running = False

        self._build_ui()
        self._check_dependencies()
        self.after(100, self._poll_queue)

    # ---------------------------------------------------------------- UI
    def _build_ui(self):
        pad = {"padx": 8, "pady": 4}

        # --- 1. Planilha de entrada
        frm_file = ttk.LabelFrame(self, text="1. Planilha (Excel / CSV)")
        frm_file.pack(fill="x", **pad)

        row = ttk.Frame(frm_file)
        row.pack(fill="x", padx=8, pady=6)
        ttk.Entry(row, textvariable=self.input_path).pack(side="left", fill="x", expand=True)
        ttk.Button(row, text="Procurar...", command=self._browse_input).pack(side="left", padx=6)

        row2 = ttk.Frame(frm_file)
        row2.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(row2, text="Coluna de e-mails (vazio = automatico):").pack(side="left")
        ttk.Entry(row2, textvariable=self.column_name, width=18).pack(side="left", padx=6)

        # --- 2. Validar (botao unico, decide pelo tamanho)
        frm_act = ttk.LabelFrame(self, text="2. Validar")
        frm_act.pack(fill="x", **pad)

        rowa = ttk.Frame(frm_act)
        rowa.pack(fill="x", padx=8, pady=6)
        self.btn_start = ttk.Button(rowa, text="Validar lista",
                                    command=self._validate_smart)
        self.btn_start.pack(side="left")
        self.btn_stop = ttk.Button(rowa, text="Parar", command=self._stop,
                                   state="disabled")
        self.btn_stop.pack(side="left", padx=6)
        self.btn_clear = ttk.Button(rowa, text="Limpar",
                                    command=self._clear_all)
        self.btn_clear.pack(side="left", padx=6)

        ttk.Checkbutton(
            frm_act, text="Remover e-mails duplicados (manter 1 por endereco) "
            "- recomendado p/ B2B",
            variable=self.dedup).pack(anchor="w", padx=8)
        ttk.Checkbutton(
            frm_act, text="Remover automaticos (noreply@, newsletter@). "
            "Caixas comerciais (contato@, comercial@) sao MANTIDAS.",
            variable=self.excluir_arriscados).pack(anchor="w", padx=8)
        ttk.Label(
            frm_act, foreground="#666",
            text=f"Arquivos com mais de {BIG_FILE_THRESHOLD:,} linhas".replace(",", ".")
            + " sao processados em streaming e gravados\ndireto num .xlsx limpo "
            "(mesmo modelo da original), sem travar a tela.\n"
            "Os menores aparecem na tabela abaixo.").pack(
            anchor="w", padx=8, pady=(2, 6))

        # --- 3. Exportar (so apos validar uma lista na tabela)
        frm_exp = ttk.LabelFrame(self, text="3. Exportar resultado da tabela")
        frm_exp.pack(fill="x", **pad)
        rowe = ttk.Frame(frm_exp)
        rowe.pack(fill="x", padx=8, pady=6)
        self.btn_clean = ttk.Button(rowe, text="Planilha LIMPA (so enviaveis)...",
                                    command=self._export_clean, state="disabled")
        self.btn_clean.pack(side="left")
        self.btn_export = ttk.Button(rowe, text="Relatorio completo...",
                                     command=self._export, state="disabled")
        self.btn_export.pack(side="left", padx=6)

        # --- Avancado (recolhivel)
        self._adv_open = False
        self.btn_adv = ttk.Button(
            self, text="▸  Avancado  (verificacao SMTP, ritmo DNS)",
            command=self._toggle_adv)
        self.btn_adv.pack(fill="x", padx=8, pady=(2, 0))
        self.frm_adv = self._build_advanced()

        # --- Progresso
        frm_prog = ttk.Frame(self)
        frm_prog.pack(fill="x", **pad)
        self.progress = ttk.Progressbar(frm_prog, mode="determinate")
        self.progress.pack(fill="x", side="left", expand=True)
        self.lbl_count = ttk.Label(frm_prog, text="0 / 0", width=14, anchor="e")
        self.lbl_count.pack(side="left", padx=6)

        # --- Resumo
        self.lbl_summary = ttk.Label(self, text="Seguros: 0   Arriscados: 0   Invalidos: 0   |   Typos corrigidos: 0")
        self.lbl_summary.pack(anchor="w", padx=10)

        # --- Tabela de resultados
        frm_tbl = ttk.LabelFrame(self, text="Resultados")
        frm_tbl.pack(fill="both", expand=True, **pad)

        cols = ("email", "status", "risco", "motivo")
        self.tree = ttk.Treeview(frm_tbl, columns=cols, show="headings")
        self.tree.heading("email", text="E-mail")
        self.tree.heading("status", text="Status")
        self.tree.heading("risco", text="Risco")
        self.tree.heading("motivo", text="Motivo")
        self.tree.column("email", width=230)
        self.tree.column("status", width=85, anchor="center")
        self.tree.column("risco", width=80, anchor="center")
        self.tree.column("motivo", width=290)
        vsb = ttk.Scrollbar(frm_tbl, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # cor por risco (mais util pro envio do que por status)
        self.tree.tag_configure("seguro", background="#e6ffe6")
        self.tree.tag_configure("arriscado", background="#fff5e6")
        self.tree.tag_configure("invalido", background="#ffe6e6")

        # --- Barra de estado
        self.status_var = tk.StringVar(value="Pronto.")
        ttk.Label(self, textvariable=self.status_var, relief="sunken",
                  anchor="w").pack(fill="x", side="bottom")

    def _build_advanced(self):
        """Painel recolhivel com os controles que a maioria nao precisa."""
        frm = ttk.LabelFrame(self, text="Avancado")

        ttk.Radiobutton(frm, text="Rapido  (sintaxe + MX do dominio) - recomendado",
                        variable=self.mode, value="rapido").pack(anchor="w", padx=8, pady=(6, 0))
        ttk.Radiobutton(frm, text="Completo  (+ SMTP) - so funciona se a porta 25 "
                        "estiver liberada (nao e o seu caso)",
                        variable=self.mode, value="completo").pack(anchor="w", padx=8, pady=(0, 2))

        row3 = ttk.Frame(frm)
        row3.pack(fill="x", padx=8, pady=2)
        ttk.Label(row3, text="Threads:").pack(side="left")
        ttk.Spinbox(row3, from_=1, to=100, textvariable=self.workers,
                    width=5).pack(side="left", padx=6)
        ttk.Label(row3, text="Ritmo SMTP (sondas/min):").pack(side="left", padx=(12, 0))
        ttk.Spinbox(row3, from_=1, to=120, textvariable=self.smtp_rate,
                    width=5).pack(side="left", padx=6)
        ttk.Label(row3, text="Ritmo DNS (consultas/seg):").pack(side="left", padx=(12, 0))
        ttk.Spinbox(row3, from_=5, to=200, textvariable=self.dns_rate,
                    width=5).pack(side="left", padx=6)

        ttk.Checkbutton(
            frm, text="Verificacao PROFUNDA (deteta catch-all + reenvia em "
            "greylisting) - mais lenta",
            variable=self.verificacao_profunda).pack(anchor="w", padx=8, pady=(2, 6))
        return frm

    def _toggle_adv(self):
        if self._adv_open:
            self.frm_adv.pack_forget()
            self.btn_adv.config(text="▸  Avancado  (verificacao SMTP, ritmo DNS)")
            self._adv_open = False
        else:
            self.frm_adv.pack(fill="x", padx=8, pady=4, after=self.btn_adv)
            self.btn_adv.config(text="▾  Avancado  (verificacao SMTP, ritmo DNS)")
            self._adv_open = True

    def _check_dependencies(self):
        missing = []
        if dns is None:
            missing.append("dnspython")
        if openpyxl is None:
            missing.append("openpyxl")
        if missing:
            messagebox.showwarning(
                "Dependencias em falta",
                "Os seguintes modulos nao estao instalados:\n  - "
                + "\n  - ".join(missing)
                + "\n\nInstale com:  pip install " + " ".join(missing))

    # ---------------------------------------------------------------- Acoes
    def _browse_input(self):
        path = filedialog.askopenfilename(
            title="Selecione a planilha",
            filetypes=[("Planilhas", "*.xlsx *.xlsm *.xls *.csv *.txt"),
                       ("Excel", "*.xlsx *.xlsm *.xls"),
                       ("CSV / Texto", "*.csv *.txt"),
                       ("Todos", "*.*")])
        if path:
            self.input_path.set(path)
            self._load_emails()

    def _load_emails(self):
        path = self.input_path.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showerror("Erro", "Selecione um ficheiro valido.")
            return
        try:
            col = self.column_name.get().strip() or None
            self.status_var.set("A carregar planilha...")
            self.update_idletasks()
            self._table = read_table_from_file(path, col)
            self._emails = self._table["emails"]
            n_linhas = len(self._table["rows"])
            n_unicos = len(set(e.strip() for e in self._emails))
            self.status_var.set(
                f"{len(self._emails)} e-mails ({n_unicos} unicos) de "
                f"{n_linhas} linhas em {os.path.basename(path)}.")
            messagebox.showinfo(
                "Carregado",
                f"Linhas de dados: {n_linhas}\n"
                f"E-mails encontrados: {len(self._emails)}\n"
                f"E-mails unicos a validar: {n_unicos}\n\n"
                f"Ao exportar a planilha limpa, todas as colunas originais "
                f"sao preservadas.")
        except Exception as exc:
            messagebox.showerror("Erro ao ler ficheiro", str(exc))
            self.status_var.set("Erro ao carregar.")

    def _validate_smart(self):
        """Botao unico: decide pelo tamanho do arquivo se vai pra tabela
        (em memoria) ou pra streaming/CSV (arquivos gigantes)."""
        if self._running:
            return
        path = self.input_path.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showwarning(
                "Sem planilha",
                "Selecione uma planilha primeiro (botao 'Procurar...').")
            return
        # descobre o tamanho rapidamente, sem carregar tudo
        try:
            _h, _ci, _rows, total = open_table_stream(
                path, self.column_name.get().strip() or None)
            _rows.close()
        except Exception as exc:
            messagebox.showerror("Erro ao abrir arquivo", str(exc))
            return

        if total and total > BIG_FILE_THRESHOLD:
            self._start_large(prefetched_total=total)
        else:
            # arquivo pequeno -> carrega e mostra na tabela
            self._load_emails()
            if self._emails:
                self._start()

    def _start(self):
        if self._running:
            return
        if not self._emails:
            self._load_emails()
            if not self._emails:
                messagebox.showwarning("Sem dados",
                                       "Carregue uma planilha com e-mails primeiro.")
                return

        # Teste de DNS antes de processar a lista inteira a toa.
        self.status_var.set("A testar conexao DNS...")
        self.update_idletasks()
        ok_dns, detalhe = dns_self_test()
        if not ok_dns:
            cont = messagebox.askyesno(
                "DNS nao esta a funcionar",
                "Nao consegui resolver DNS para um dominio conhecido.\n\n"
                f"Detalhe: {detalhe}\n\n"
                "Sem DNS, a verificacao de MX falha e os e-mails ficarao como "
                "'desconhecido' (nao serao marcados como invalidos por engano).\n\n"
                "Verifique a sua internet/firewall. Deseja continuar mesmo assim?")
            if not cont:
                self.status_var.set("Cancelado: DNS indisponivel.")
                return

        # valida apenas e-mails unicos (mais rapido); o resultado e depois
        # aplicado a todas as linhas pela coluna de e-mail.
        uniq = list(dict.fromkeys(e.strip() for e in self._emails))

        mode = self.mode.get()
        n_workers = max(1, self.workers.get())
        deep = self.verificacao_profunda.get()

        if mode == "completo":
            # --- SEGURANCAS ANTI-BLOQUEIO ---
            # 1) limita as threads automaticamente
            n_workers = min(n_workers, SMTP_MAX_WORKERS)
            # 2) define o ritmo global de sondas
            rate = max(1, self.smtp_rate.get())
            set_smtp_rate_per_min(rate)
            # 3) estima quantas sondas SMTP de facto serao feitas (exclui os
            #    provedores accept-all, que sao aceites sem sondar)
            sondaveis = sum(
                1 for e in uniq
                if "@" in e and e.rsplit("@", 1)[-1].lower() not in ACCEPT_ALL_DOMAINS)
            # na verificacao profunda ha +1 sonda (catch-all) por dominio unico
            dominios_unicos = len({
                e.rsplit("@", 1)[-1].lower() for e in uniq
                if "@" in e and e.rsplit("@", 1)[-1].lower() not in ACCEPT_ALL_DOMAINS})
            total_sondas = sondaveis + (dominios_unicos if deep else 0)
            minutos = total_sondas / rate if rate else 0
            tempo = (f"{minutos:.0f} min" if minutos < 60
                     else f"{minutos/60:.1f} h")

            if deep:
                linha_profunda = (
                    f"  - PROFUNDA: deteta catch-all (+{dominios_unicos} sondas) "
                    f"e reenvia em greylisting (espera {SMTP_GREYLIST_DELAY}s, "
                    f"ate {SMTP_GREYLIST_RETRIES}x)\n")
            else:
                linha_profunda = "  - Profunda: DESATIVADA (sonda simples)\n"

            ok = messagebox.askyesno(
                "Modo completo (SMTP) - seguranca ativa",
                "A verificacao SMTP sera feita DE FORMA GRADUAL para evitar "
                "bloqueio/blacklist do seu IP:\n\n"
                f"  - Threads limitadas a {n_workers}\n"
                f"  - Ritmo: {rate} sondas/min\n"
                f"  - 1 sonda por vez por dominio (intervalo de "
                f"{SMTP_PER_DOMAIN_INTERVAL:.0f}s)\n"
                f"  - Gmail/Outlook/Yahoo e afins: aceites pelo MX, sem sondar\n"
                + linha_profunda +
                f"\nE-mails a sondar via SMTP: ~{sondaveis}\n"
                f"Sondas totais estimadas: ~{total_sondas}\n"
                f"Tempo estimado: ~{tempo} (greylisting pode aumentar)\n\n"
                "Obs.: muitos ISPs bloqueiam a porta 25 (resultados ficam "
                "'desconhecido'). Deseja continuar?")
            if not ok:
                return

        # limpa estado
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._results = []
        self._stop_flag.clear()
        self._running = True
        self.progress["value"] = 0
        self.progress["maximum"] = len(uniq)
        self.lbl_count.config(text=f"0 / {len(uniq)}")
        self._update_summary()

        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.btn_export.config(state="disabled")
        self.btn_clean.config(state="disabled")
        self.status_var.set(
            f"A validar {len(uniq)} e-mails unicos em modo {mode} "
            f"({n_workers} threads)...")

        worker = threading.Thread(
            target=self._run_validation,
            args=(uniq, mode, n_workers, deep),
            daemon=True)
        worker.start()

    def _run_validation(self, emails, mode, n_workers, deep=False):
        """Executa em thread de fundo; comunica via fila."""
        try:
            with ThreadPoolExecutor(max_workers=n_workers) as ex:
                futures = {ex.submit(validate_email, e, mode, deep): e
                           for e in emails}
                for fut in as_completed(futures):
                    if self._stop_flag.is_set():
                        break
                    try:
                        res = fut.result()
                    except Exception as exc:
                        res = {"email": futures[fut], "status": "invalido",
                               "motivo": f"Erro: {exc}"}
                    self._queue.put(("result", res))
        finally:
            self._queue.put(("done", None))

    def _poll_queue(self):
        """Drena a fila e atualiza a GUI (corre na thread principal)."""
        try:
            while True:
                kind, payload = self._queue.get_nowait()
                if kind == "result":
                    self._add_result(payload)
                elif kind == "done":
                    self._on_done()
                elif kind == "big_progress":
                    self._big_update(payload)
                elif kind == "big_done":
                    self._big_done(payload)
                elif kind == "big_error":
                    self._big_error(payload)
        except queue.Empty:
            pass
        self.after(80, self._poll_queue)

    def _add_result(self, res):
        self._results.append(res)
        mostra = res.get("email_final") or res.get("email", "")
        self.tree.insert("", "end",
                         values=(mostra, res["status"],
                                 res.get("risco", ""), res["motivo"]),
                         tags=(res.get("risco", res["status"]),))
        n = len(self._results)
        self.progress["value"] = n
        self.lbl_count.config(text=f"{n} / {self.progress['maximum']}")
        # auto-scroll
        children = self.tree.get_children()
        if children:
            self.tree.see(children[-1])
        if n % 25 == 0:
            self._update_summary()

    def _on_done(self):
        if not self._running:
            return
        self._running = False
        self._update_summary()
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        has = bool(self._results)
        self.btn_export.config(state="normal" if has else "disabled")
        self.btn_clean.config(state="normal" if has and self._table else "disabled")
        stopped = self._stop_flag.is_set()
        self.status_var.set("Validacao interrompida." if stopped
                            else "Validacao concluida.")

    def _update_summary(self):
        seg = sum(1 for r in self._results if r.get("risco") == "seguro")
        arr = sum(1 for r in self._results if r.get("risco") == "arriscado")
        inv = sum(1 for r in self._results if r.get("risco") == "invalido")
        corr = sum(1 for r in self._results if r.get("corrigido"))
        self.lbl_summary.config(
            text=f"Seguros: {seg}   Arriscados: {arr}   Invalidos: {inv}"
                 f"   |   Typos corrigidos: {corr}")

    def _clear_all(self):
        if self._running:
            messagebox.showinfo(
                "Ocupado",
                "Ha uma tarefa em andamento. Clique em 'Parar' antes de limpar.")
            return
        # limpa tabela e todo o estado, deixando pronto para outra planilha
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._results = []
        self._emails = []
        self._table = None
        self.input_path.set("")
        self.column_name.set("")
        self.progress["value"] = 0
        self.progress["maximum"] = 100
        self.lbl_count.config(text="0 / 0")
        self.lbl_summary.config(
            text="Seguros: 0   Arriscados: 0   Invalidos: 0"
                 "   |   Typos corrigidos: 0")
        self.btn_export.config(state="disabled")
        self.btn_clean.config(state="disabled")
        self.status_var.set("Limpo. Carregue outra planilha.")

    def _stop(self):
        if self._running:
            self._stop_flag.set()
            self.status_var.set("A parar...")

    # ------------------------------------------------ Grande escala (streaming)
    def _start_large(self, prefetched_total=None):
        if self._running:
            messagebox.showinfo("Ocupado", "Aguarde terminar a tarefa atual.")
            return
        path = self.input_path.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showwarning(
                "Sem planilha",
                "Selecione uma planilha primeiro (botao 'Procurar...').")
            return

        # estima rapidamente o tamanho (sem carregar tudo), salvo se ja sabemos
        total = prefetched_total
        if total is None:
            try:
                _h, _ci, _rows, total = open_table_stream(
                    path, self.column_name.get().strip() or None)
                _rows.close()  # fecha o handle sem consumir o gerador
            except Exception as exc:
                messagebox.showerror("Erro ao abrir arquivo", str(exc))
                return

        rate = max(1, self.dns_rate.get())
        excluir = self.excluir_arriscados.get()
        dedup = self.dedup.get()
        est_txt = f"~{total:,} linhas".replace(",", ".") if total else "desconhecido"

        ok_dns, det = dns_self_test()
        if not ok_dns:
            if not messagebox.askyesno(
                    "DNS nao esta a funcionar",
                    f"Sem DNS a validacao por MX falha.\n\nDetalhe: {det}\n\n"
                    "Continuar mesmo assim?"):
                return

        if not messagebox.askyesno(
                "Arquivo grande -> Excel em streaming",
                f"Arquivo: {os.path.basename(path)}\n"
                f"Linhas estimadas: {est_txt}  (acima de "
                f"{BIG_FILE_THRESHOLD:,}".replace(",", ".") + ")\n"
                f"Ritmo DNS: {rate} consultas/seg\n"
                f"Duplicados: {'removidos (1 por endereco)' if dedup else 'mantidos'}\n"
                f"Automaticos (noreply/newsletter): "
                f"{'removidos' if excluir else 'mantidos'}\n\n"
                "Por ser grande, vou process-lo em STREAMING (modo Rapido/MX), "
                "sem travar a tela, e gravar uma planilha .xlsx limpa com o "
                "MESMO modelo da original (todas as colunas e cabecalho), "
                "e-mails com typo corrigidos, so linhas enviaveis.\n\n"
                "A seguir escolha onde salvar. Continuar?"):
            return

        sugest = os.path.splitext(path)[0] + "_LIMPA.xlsx"
        out = filedialog.asksaveasfilename(
            title="Salvar planilha limpa (Excel)",
            initialfile=os.path.basename(sugest),
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")])
        if not out:
            return

        # prepara estado
        self._running = True
        self._stop_flag.clear()
        self._big_t0 = time.monotonic()
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.progress["value"] = 0
        self.progress["maximum"] = total if total else 100
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.btn_export.config(state="disabled")
        self.btn_clean.config(state="disabled")
        self.status_var.set("Processando em grande escala...")

        col = self.column_name.get().strip() or None
        worker = threading.Thread(
            target=self._run_large,
            args=(path, out, col, excluir, rate, dedup),
            daemon=True)
        worker.start()

    def _run_large(self, in_path, out_path, col, excluir, rate, dedup=False):
        """Thread de fundo: processa o arquivo grande em streaming."""
        def progress_cb(counters):
            counters["_out"] = out_path
            self._queue.put(("big_progress", counters))
        try:
            final = stream_process(
                in_path, out_path, col, excluir, rate,
                progress_cb=progress_cb, stop_event=self._stop_flag,
                dedup=dedup)
            final["_out"] = out_path
            self._queue.put(("big_done", final))
        except Exception as exc:
            self._queue.put(("big_error", str(exc)))

    def _big_update(self, c):
        proc = c.get("processadas", 0)
        total = c.get("total_estimado", 0)
        self.progress["maximum"] = total if total else max(proc, 1)
        self.progress["value"] = min(proc, self.progress["maximum"])
        elapsed = max(0.001, time.monotonic() - self._big_t0)
        rate = proc / elapsed
        eta = ""
        if total and rate > 0 and proc < total:
            secs = (total - proc) / rate
            eta = (f"  ETA ~{secs/60:.0f} min" if secs >= 60
                   else f"  ETA ~{secs:.0f} s")
        tot_txt = f"{total:,}".replace(",", ".") if total else "?"
        self.lbl_count.config(
            text=f"{proc:,}".replace(",", ".") + f" / {tot_txt}")
        self.lbl_summary.config(
            text=f"Seguros: {c.get('seguro',0):,}".replace(",", ".")
                 + f"   Arriscados: {c.get('arriscado',0):,}".replace(",", ".")
                 + f"   Invalidos: {c.get('invalido',0):,}".replace(",", ".")
                 + f"   Sem e-mail: {c.get('sem_email',0):,}".replace(",", ".")
                 + f"   |   Mantidas: {c.get('mantidas',0):,}".replace(",", ".")
                 + f"   Dupl.: {c.get('duplicados',0):,}".replace(",", ".")
                 + f"   Typos: {c.get('corrigido',0):,}".replace(",", "."))
        self.status_var.set(
            f"Processando... {rate:.0f} linhas/s{eta}")

    def _big_done(self, c):
        self._running = False
        self._big_update(c)
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        stopped = self._stop_flag.is_set()
        out = c.get("_out", "")
        elapsed = time.monotonic() - self._big_t0
        self.status_var.set(
            ("Interrompido. " if stopped else "Concluido. ")
            + f"Planilha limpa: {os.path.basename(out)}")
        messagebox.showinfo(
            "Interrompido" if stopped else "Concluido",
            ("PARCIAL (interrompido)\n\n" if stopped else "")
            + f"Arquivo limpo:\n{out}\n\n"
            f"Processadas : {c.get('processadas',0):,}".replace(",", ".") + "\n"
            f"Sem e-mail  : {c.get('sem_email',0):,}".replace(",", ".") + "\n"
            f"Seguros     : {c.get('seguro',0):,}".replace(",", ".") + "\n"
            f"Arriscados  : {c.get('arriscado',0):,}".replace(",", ".") + "\n"
            f"Invalidos   : {c.get('invalido',0):,}".replace(",", ".") + "\n"
            f"Typos corr. : {c.get('corrigido',0):,}".replace(",", ".") + "\n"
            f"Duplicados  : {c.get('duplicados',0):,}".replace(",", ".") + "\n"
            f"------------------------------\n"
            f"MANTIDAS na planilha: {c.get('mantidas',0):,}".replace(",", ".") + "\n\n"
            f"Tempo: {elapsed/60:.1f} min")

    def _big_error(self, msg):
        self._running = False
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.status_var.set("Erro no processamento.")
        messagebox.showerror("Erro ao processar", msg)

    def _export(self):
        if not self._results:
            return
        path = filedialog.asksaveasfilename(
            title="Guardar resultados",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")])
        if not path:
            return
        try:
            write_results(path, self._results)
            self.status_var.set(f"Relatorio guardado em {os.path.basename(path)}.")
            messagebox.showinfo("Exportado", f"Relatorio guardado em:\n{path}")
        except Exception as exc:
            messagebox.showerror("Erro ao exportar", str(exc))

    def _export_clean(self):
        """Exporta a planilha original mantendo apenas os e-mails enviaveis."""
        if not self._results or not self._table:
            return
        path = filedialog.asksaveasfilename(
            title="Guardar planilha limpa",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")])
        if not path:
            return

        # mapa email original -> result completo (normalizado igual a leitura)
        info_by_email = {r["email"].strip(): r for r in self._results}
        excluir_arr = self.excluir_arriscados.get()

        def keep_fn(info):
            if info.get("status") != "valido":
                return False
            if excluir_arr and info.get("risco") == "arriscado":
                return False
            return True

        try:
            total, kept = write_clean_spreadsheet(
                path, self._table, keep_fn, info_by_email,
                dedup=self.dedup.get())
            removed = total - kept
            extra = ("Arriscados (e-mails de funcao) tambem foram removidos.\n"
                     if excluir_arr else
                     "E-mails de funcao (arriscados) foram MANTIDOS.\n")
            self.status_var.set(
                f"Planilha limpa: {kept} mantidas, {removed} removidas.")
            messagebox.showinfo(
                "Planilha limpa exportada",
                f"Arquivo: {path}\n\n"
                f"Linhas originais : {total}\n"
                f"Mantidas          : {kept}\n"
                f"Removidas         : {removed}\n\n"
                + extra +
                "Os e-mails com typo foram corrigidos na saida.\n"
                "Todas as colunas originais foram preservadas.")
        except Exception as exc:
            messagebox.showerror("Erro ao exportar", str(exc))


def main():
    app = EmailValidatorApp()
    app.mainloop()


if __name__ == "__main__":
    main()
